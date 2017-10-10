## testfile

import openpyxl
import sys
import math
import logging
import datetime
import NXOpen
import NXOpen.Preferences
import os

## general settings
overwrite = True

## excel file variables
excel_file = "SimpleParameters.xlsx"
excel_path = "C:\\Users\\mujakob\\Box Sync\\MEPHISTO\\EWB_test\\TESTING"
excel_fullpath = os.path.join(excel_path, excel_file)
excel_sheet_name = "Sheet1"
design_case_indicator_column='DC' # name (word in row one) of the column containig the design case identifier

## NX file variables
master_cad_file = "MikroTRS.prt"
master_cad_path = "C:\\Users\\mujakob\\Box Sync\\MEPHISTO\\EWB_test\\TESTING"
master_cad_fullpath = os.path.join(master_cad_path, master_cad_file)

#NX output
final_path = "C:\\Users\\mujakob\\Box Sync\\MEPHISTO\\EWB_test\\TESTING\\Results1"
final_name = "SimpleTRS"
#NX globals
theSession  = NXOpen.Session.GetSession()
workPart = NXOpen.Part.Null
displayPart = NXOpen.Part.Null

######################################################

def EndWithError(reason):
    logging.error("Premature Termination because of {0}".format(reason))
    sys.exit(reason)

def checkForDirOrCreate(directory):
    #checks if "directory" exists, and if not, tries to create it. if it fails, kills all
    if os.path.isdir(directory):
        logging.debug("Directory {0} exists - Good!".format(directory))
        #return True
    else:
        try:
            os.makedirs(directory)
            logging.info("Created {0}".format(directory))
        except:
            EndWithError("Could neither find nor create {0}".format(directory))
            
def checkForFileOrDie(file):
    # checks if "file" exists, and kills the script if so
    if os.path.exists(file):
        logging.debug("File {0} exists - Good!".format(file))
        return True
    else:
        EndWithError("Could not find {0}".format(file))
        
def overwriteFile(file):
    # checks if "file" exists, and deletes it if "overwrite" is set accordingly. returns true if file is gone (or has never been)
    if os.path.exists(file):
        if overwrite:
            try:
                os.remove(file)
                logging.info("Deleted {0} because overwrite is set to 'True'".format(file))
                return True()
            except:
                logging.warning("Could not delete {0}, although overwrite is set to 'True'".format(file))
                return(False)
        else:
            logging.info("Did _not_ overwrite {0}, debug is 'False'".format(file))
            return False
    else:
        logging.debug("No need to overwrite {0}, it doesn't exist".format(file))
        return True
        
def overwriteFileOrRename(file):
    # iterates file names until one is found that does not exist, returns virgin file name
    f_count = 0
    while not overwriteFile(file):
        f_name, f_ext = os.path.splitext(file)
        if f_count == 0:
            f_count += 1
            f_name = f_name + "_" + str(f_count)
        else:
            f_name = f_name.rstrip("0123456789") + f_count
        file = f_name + f_ext
    return file
    
def TakeApicture(wp, imageName):
    printPDFBuilder1 = wp.PlotManager.CreatePrintPdfbuilder()
    
    printPDFBuilder1.Scale = 1.0
    
    printPDFBuilder1.Size = NXOpen.PrintPDFBuilder.SizeOption.ScaleFactor
    
    printPDFBuilder1.Units = NXOpen.PrintPDFBuilder.UnitsOption.English
    
    printPDFBuilder1.XDimension = 8.5
    
    printPDFBuilder1.YDimension = 11.0
    
    printPDFBuilder1.RasterImages = True
    
    printPDFBuilder1.OutputText = NXOpen.PrintPDFBuilder.OutputTextOption.Text
    
    printPDFBuilder1.ImageResolution = NXOpen.PrintPDFBuilder.ImageResolutionOption.Medium

    printPDFBuilder1.Watermark = "VITUM"
    
    sheets1 = [NXOpen.NXObject.Null] * 1 
    sheets1[0] = NXOpen.NXObject.Null
    printPDFBuilder1.SourceBuilder.SetSheets(sheets1)
    
    printPDFBuilder1.Filename = imageName
    
    nXObject1 = printPDFBuilder1.Commit()
    
    printPDFBuilder1.Destroy()

    # # NX-part class to be developed!
class NX_part:
    ''' a class to handle all NX part related operations
    requires a running NX session and running logger'''
    
    def __init__(self, session, cad_file):
        ''' open the file and set workpart and displaypart '''
        self.session = session
        if os.path.exists(cad_file):
                
            basePart1, partLoadStatus1 = theSession.Parts.OpenBaseDisplay(cad_file)
            self.workPart = session.Parts.Work
            self.displayPart = session.Parts.Display
            
            logging.debug("Created new NX object for {0}".format(cad_file))
        else:
            logging.error("Could not open {0}".format(cad_file))
    
    def ChangeParameter(self, pName, pValue, holdUpdate = False):
        ''' takes a parameter name and value, searches the part for it and changes the respective expression, 
        taking into account the expression type of the part, -- to be developed '''
        
        #placing an undo mark in case the update fails
        markId = theSession.SetUndoMark(NXOpen.Session.MarkVisibility.Invisible, "NX update")
        
        # gathering info about the expression:
        try:
            expression = self.workPart.Expressions.FindObject(pName)
        except:
            logging.info("Could not find expression {0}".format(pName))
            return
        type = expression.Type
        logging.debug("attemting to change expression {0} of type {1} to value {2} ".format(pName, type, pValue))
        

        if type == "Number":
            units = expression.Units
            pValue = float(pValue)
        else:
            units = None
        
        # now we go and set it:
        if units != None: 
            self.workPart.Expressions.EditWithUnits(expression, units, str(pValue))   # edit of number
        else:
            self.workPart.Expressions.Edit(expression, str(pValue))   # edit of number
        
        
        if not holdUpdate:
            nErrs1 = self.session.UpdateManager.DoUpdate(markId)
            logging.debug("changed {0} to {1} and Updated, erros: {2} ".format(pName, pValue, str(nErrs1)))
        else:
            logging.debug("changed {0} to {1} and WITHOUT update ".format(pName, pValue))    
        
    def TakePicture(self, iName):
        ''' exports a PDF to the os.path and name given in iName '''
        printPDFBuilder1 = self.workPart.PlotManager.CreatePrintPdfbuilder()
    
        printPDFBuilder1.Scale = 1.0
        
        printPDFBuilder1.Size = NXOpen.PrintPDFBuilder.SizeOption.ScaleFactor
        
        printPDFBuilder1.Units = NXOpen.PrintPDFBuilder.UnitsOption.English
        
        printPDFBuilder1.XDimension = 8.5
        
        printPDFBuilder1.YDimension = 11.0
        
        printPDFBuilder1.RasterImages = True
        
        printPDFBuilder1.OutputText = NXOpen.PrintPDFBuilder.OutputTextOption.Text
        
        printPDFBuilder1.ImageResolution = NXOpen.PrintPDFBuilder.ImageResolutionOption.Medium

        printPDFBuilder1.Watermark = "VITUM"
        
        sheets1 = [NXOpen.NXObject.Null] * 1 
        sheets1[0] = NXOpen.NXObject.Null
        printPDFBuilder1.SourceBuilder.SetSheets(sheets1)
        
        printPDFBuilder1.Filename = iName
        
        nXObject1 = printPDFBuilder1.Commit()
        
        printPDFBuilder1.Destroy()

    def Save(self, fName):
        ''' saves the part to the os.path fPath '''
    
        final_fullpath, ext = os.path.splitext(fName) # of course we also need to remove the ext again ... 
        
        partSaveStatus1 = self.workPart.SaveAs(final_fullpath)
        
        partSaveStatus1.Dispose()
        logging.info("saved part as {0}".format(fName))
        
    def Update(self):
        # # Update model ###
        markId6 = self.session.SetUndoMark(NXOpen.Session.MarkVisibility.Invisible, "NX update")
           
        nErrs1 = self.session.UpdateManager.DoUpdate(markId6)
        
    def Close(self):
        self.workPart.Close(NXOpen.BasePart.CloseWholeTree.TrueValue, NXOpen.BasePart.CloseModified.UseResponses, None)
     
     
     
def main():
    
    # # logging setup
    if not os.path.isdir(final_path):
        try:
            os.makedirs(final_path)
            logging.info("Created {0}".format(final_path))
        except:
            exit("Could not Create Target Directory!")
    log_file = "test.txt" #str(datetime.datetime.now())+"_GeoVar_Logging.txt"
    log_fullpath = os.path.join(final_path, log_file)
    logging.basicConfig(filename=log_fullpath, filemode='w', level=logging.DEBUG)
    logging.info("##### Good Morning, this is a new run! ##### ")
    
    # # Input validation
    checkForFileOrDie(excel_fullpath)
    checkForFileOrDie(master_cad_fullpath)
    checkForDirOrCreate(final_path)
    
    # # Open Excel, read parameter ranges and such
    wb = openpyxl.load_workbook(excel_fullpath, data_only=True)
    sheets_in_wb=(wb.get_sheet_names())
    # check for worksheet name in file:
    if not excel_sheet_name in sheets_in_wb:
        EndWithError('Could not find Excel sheet: %s' % {excel_sheet_name})
        
    # open WS
    ws = wb[excel_sheet_name]
    maxDS = ws.max_row - 1
    InputRange = "A1:{0}{1}".format(openpyxl.utils.get_column_letter(ws.max_column), ws.max_row)
    
    logging.info("found {0} design cases in {1}".format(maxDS, InputRange))
    
    # create parameter dictionary
    # parameter = [ {[p1_Name:p1DC1_value]; [p2_Name:p1DC2_value]; ... 
    #                        {[p1_Name:p1DC2_value]; ...} 
    #                      ]
    parameters = [{n:v for (n,v) in zip((c1.value for c1 in ws[1]), (cn.value for cn in ws[n]))} for n in range(2, ws.max_row)]
    
    for p in parameters:
        logging.debug("{0} /n".format(p))

    # #####################################################    
    # # NX
    theSession  = NXOpen.Session.GetSession()
    workPart = theSession.Parts.Work
    displayPart = theSession.Parts.Display
    theSession.Preferences.Modeling.UpdatePending = False
    
    DC = 1
   
    #while DC <= maxDS:
    for p in parameters:
        try:
            DC = p[design_case_indicator_column]
        except:
            logging.error('no designcase found! skipping...')
            continue
         
        logging.info("Starting with DC {0}".format(DC))
        # # Open master file
        nxPart = NX_part(theSession, master_cad_fullpath)
       
        for name, value in p.items():
            nxPart.ChangeParameter(name, value, True)
        
        nxPart.Update()
        # # Save CAD
        final_name_DC = "{0}{1}".format(final_name, DC)
        final_fullpath = os.path.join(final_path, final_name_DC)
        final_fullpath = overwriteFileOrRename(final_fullpath + ".prt") # need to add extension but it is not needed for the NX save procedure
        nxPart.Save(final_fullpath)
        
        image_name = "img_{0}.pdf".format(final_name_DC)
        image_fullpath = os.path.join(final_path, image_name)
        image_fullpath = overwriteFileOrRename(image_fullpath)
        nxPart.TakePicture(image_fullpath)
       
        nxPart.Close()

        workPart = NXOpen.Part.Null
        displayPart = NXOpen.Part.Null
        theSession.ApplicationSwitchImmediate("UG_APP_NOPART")
        
        #ready for next round!

    

if __name__ == "__main__":
    main()
